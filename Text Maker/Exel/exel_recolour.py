from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def highlight_asterisk_cells(input_file, output_file):
    # Загружаем книгу
    wb = load_workbook(input_file)

    # Цвет заливки (желтый)
    fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and "*" in cell.value:
                    cell.fill = fill

    # Сохраняем в новый файл
    wb.save(output_file)
    print(f"Обработанный файл сохранён как {output_file}")



# Пример использования
input_path = ['/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииВСЕХ.xlsx',
              '/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииПсихологи.xlsx',
              '/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииВрачи.xlsx']
output_path = ["/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииВСЕХ_UPD.xlsx",
               "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииПсихологи_UPD.xlsx",
               "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииВрачи_UPD.xlsx"]

for i in range(len(input_path)):
    highlight_asterisk_cells(input_path[i], output_path[i])

