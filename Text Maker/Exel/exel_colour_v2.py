import logging
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from tqdm import tqdm

# Настройка логирования
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[logging.StreamHandler()]  # вывод в консоль
)

def highlight_asterisk_cells(input_file, output_file):
    try:
        logging.info(f"Обрабатываю файл: {input_file}\n")
        wb = load_workbook(input_file)
        fill = PatternFill(start_color="008000", end_color="008000", fill_type="solid")

        for sheet in wb.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str) and "***" in cell.value:
                        cell.fill = fill

        # Создаём папку для вывода, если её нет
        os.makedirs(os.path.dirname(output_file), exist_ok=True)

        wb.save(output_file)
        logging.info(f"Файл сохранён: {output_file}\n")
    except Exception as e:
        logging.error(f"Ошибка при обработке файла {input_file}: {e}")

if __name__ == "__main__":
    input_path = [
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииВСЕХ_UPD_v2.xlsx",
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииПсихологи_UPD_v2.xlsx",
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/files_exel/КорреляцииВрачи_UPD_v2.xlsx"
    ]

    output_path = [
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииВСЕХ_UPD_v3.xlsx",
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииПсихологи_UPD_v3.xlsx",
        "/Users/antonkuzmichev/VKR_researching/VKR_research_repo/Text Maker/Exel/colored_files_exel/КорреляцииВрачи_UPD_v3.xlsx"
    ]

    # Используем tqdm для отображения прогресса
    for inp, outp in tqdm(zip(input_path, output_path), total=len(input_path), desc="Обработка файлов"):
        highlight_asterisk_cells(inp, outp)
